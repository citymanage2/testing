import uuid
from datetime import date, datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, status, Body
from pydantic import BaseModel
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.auth import get_current_user
from app.models.user import User
from app.models.task import Task
from app.models.estimate_item import EstimateItem
from app.models.contractor import Contractor
from app.models.work_acceptance import SubcontractorAssignment, WorkAcceptance, WorkAcceptanceItem
from app.models.subcontractor_contract import SubcontractorContract, SubcontractorContractItem

router = APIRouter(tags=["work-acceptances"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_estimate(task_id: str, user: User, db: AsyncSession) -> Task:
    """Fetch the task (estimate) and verify ownership."""
    result = await db.execute(select(Task).where(Task.id == task_id))
    task = result.scalar_one_or_none()
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Estimate not found")
    if task.user_id != user.id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Access denied")
    return task


async def _get_acceptance(acc_id: str, task_id: str, db: AsyncSession) -> WorkAcceptance:
    result = await db.execute(
        select(WorkAcceptance).where(
            WorkAcceptance.id == acc_id,
            WorkAcceptance.estimate_id == task_id,
        )
    )
    acc = result.scalar_one_or_none()
    if acc is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Acceptance not found")
    return acc


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class AssignmentCreate(BaseModel):
    contractor_id: str
    scope_type: str = "section"
    scope_ref: Optional[str] = None
    notes: Optional[str] = None


class AssignmentOut(BaseModel):
    id: str
    estimate_id: str
    contractor_id: str
    scope_type: str
    scope_ref: Optional[str]
    notes: Optional[str]
    created_at: datetime

    class Config:
        from_attributes = True


class AcceptanceCreate(BaseModel):
    contractor_id: Optional[str] = None
    act_number: str = "1"
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    notes: Optional[str] = None


class AcceptancePatch(BaseModel):
    status: Optional[str] = None
    notes: Optional[str] = None
    act_number: Optional[str] = None
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    contractor_id: Optional[str] = None


class AcceptanceOut(BaseModel):
    id: str
    estimate_id: str
    contractor_id: Optional[str]
    act_number: str
    period_start: Optional[date]
    period_end: Optional[date]
    notes: Optional[str]
    status: str
    created_at: datetime
    items_count: int = 0
    total_accepted_value: float = 0.0

    class Config:
        from_attributes = True


class AcceptanceItemIn(BaseModel):
    estimate_item_id: str
    quantity_accepted: float


class AcceptanceItemOut(BaseModel):
    id: str
    acceptance_id: str
    estimate_item_id: str
    quantity_accepted: float

    class Config:
        from_attributes = True


class AcceptanceProgressItem(BaseModel):
    item_id: str
    item_name: str
    unit: Optional[str]
    quantity_total: float
    quantity_accepted: float
    quantity_remaining: float
    pct_complete: float


# ---------------------------------------------------------------------------
# Subcontractor assignments
# ---------------------------------------------------------------------------

@router.get("/estimates/{task_id}/assignments", response_model=List[AssignmentOut])
async def list_assignments(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)
    result = await db.execute(
        select(SubcontractorAssignment)
        .where(SubcontractorAssignment.estimate_id == task_id)
        .order_by(SubcontractorAssignment.created_at)
    )
    return result.scalars().all()


@router.post("/estimates/{task_id}/assignments", response_model=AssignmentOut, status_code=status.HTTP_201_CREATED)
async def create_assignment(
    task_id: str,
    body: AssignmentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)

    # Verify contractor exists
    contractor_result = await db.execute(select(Contractor).where(Contractor.id == body.contractor_id))
    if contractor_result.scalar_one_or_none() is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contractor not found")

    assignment = SubcontractorAssignment(
        id=str(uuid.uuid4()),
        estimate_id=task_id,
        contractor_id=body.contractor_id,
        scope_type=body.scope_type,
        scope_ref=body.scope_ref,
        notes=body.notes,
        created_at=datetime.now(timezone.utc),
    )
    db.add(assignment)
    await db.commit()
    await db.refresh(assignment)
    return assignment


@router.delete("/estimates/{task_id}/assignments/{asgn_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_assignment(
    task_id: str,
    asgn_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)
    result = await db.execute(
        select(SubcontractorAssignment).where(
            SubcontractorAssignment.id == asgn_id,
            SubcontractorAssignment.estimate_id == task_id,
        )
    )
    assignment = result.scalar_one_or_none()
    if assignment is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Assignment not found")
    await db.delete(assignment)
    await db.commit()


# ---------------------------------------------------------------------------
# Work acceptances
# ---------------------------------------------------------------------------

@router.get("/estimates/{task_id}/acceptances", response_model=List[AcceptanceOut])
async def list_acceptances(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)

    acc_result = await db.execute(
        select(WorkAcceptance)
        .where(WorkAcceptance.estimate_id == task_id)
        .order_by(WorkAcceptance.created_at)
    )
    acceptances = acc_result.scalars().all()

    # Enrich with items_count and total_accepted_value
    output: List[AcceptanceOut] = []
    for acc in acceptances:
        # Count items
        count_result = await db.execute(
            select(func.count(WorkAcceptanceItem.id)).where(
                WorkAcceptanceItem.acceptance_id == acc.id
            )
        )
        items_count = count_result.scalar_one() or 0

        # Total accepted value = sum(quantity_accepted * (work_price + mat_price)) joined with estimate_items
        value_result = await db.execute(
            select(func.coalesce(func.sum(WorkAcceptanceItem.quantity_accepted * (EstimateItem.work_price + EstimateItem.mat_price)), 0.0))
            .join(EstimateItem, WorkAcceptanceItem.estimate_item_id == EstimateItem.id)
            .where(WorkAcceptanceItem.acceptance_id == acc.id)
        )
        total_value = value_result.scalar_one() or 0.0

        out = AcceptanceOut(
            id=acc.id,
            estimate_id=acc.estimate_id,
            contractor_id=acc.contractor_id,
            act_number=acc.act_number,
            period_start=acc.period_start,
            period_end=acc.period_end,
            notes=acc.notes,
            status=acc.status,
            created_at=acc.created_at,
            items_count=items_count,
            total_accepted_value=total_value,
        )
        output.append(out)

    return output


@router.post("/estimates/{task_id}/acceptances", response_model=AcceptanceOut, status_code=status.HTTP_201_CREATED)
async def create_acceptance(
    task_id: str,
    body: AcceptanceCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)

    if body.contractor_id is not None:
        contractor_result = await db.execute(select(Contractor).where(Contractor.id == body.contractor_id))
        if contractor_result.scalar_one_or_none() is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contractor not found")

    acceptance = WorkAcceptance(
        id=str(uuid.uuid4()),
        estimate_id=task_id,
        contractor_id=body.contractor_id,
        act_number=body.act_number,
        period_start=body.period_start,
        period_end=body.period_end,
        notes=body.notes,
        status="draft",
        created_at=datetime.now(timezone.utc),
    )
    db.add(acceptance)
    await db.commit()
    await db.refresh(acceptance)

    return AcceptanceOut(
        id=acceptance.id,
        estimate_id=acceptance.estimate_id,
        contractor_id=acceptance.contractor_id,
        act_number=acceptance.act_number,
        period_start=acceptance.period_start,
        period_end=acceptance.period_end,
        notes=acceptance.notes,
        status=acceptance.status,
        created_at=acceptance.created_at,
        items_count=0,
        total_accepted_value=0.0,
    )


@router.patch("/estimates/{task_id}/acceptances/{acc_id}", response_model=AcceptanceOut)
async def patch_acceptance(
    task_id: str,
    acc_id: str,
    body: AcceptancePatch,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)
    acc = await _get_acceptance(acc_id, task_id, db)

    if body.status is not None:
        allowed_statuses = {"draft", "accepted", "rejected"}
        if body.status not in allowed_statuses:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Invalid status. Must be one of: {', '.join(allowed_statuses)}",
            )
        acc.status = body.status

    if body.notes is not None:
        acc.notes = body.notes

    if body.act_number is not None:
        acc.act_number = body.act_number

    if body.period_start is not None:
        acc.period_start = body.period_start

    if body.period_end is not None:
        acc.period_end = body.period_end

    if body.contractor_id is not None:
        contractor_result = await db.execute(select(Contractor).where(Contractor.id == body.contractor_id))
        if contractor_result.scalar_one_or_none() is None:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Contractor not found")
        acc.contractor_id = body.contractor_id

    await db.commit()
    await db.refresh(acc)

    # Compute enrichment fields
    count_result = await db.execute(
        select(func.count(WorkAcceptanceItem.id)).where(WorkAcceptanceItem.acceptance_id == acc.id)
    )
    items_count = count_result.scalar_one() or 0

    value_result = await db.execute(
        select(func.coalesce(func.sum(WorkAcceptanceItem.quantity_accepted * (EstimateItem.work_price + EstimateItem.mat_price)), 0.0))
        .join(EstimateItem, WorkAcceptanceItem.estimate_item_id == EstimateItem.id)
        .where(WorkAcceptanceItem.acceptance_id == acc.id)
    )
    total_value = value_result.scalar_one() or 0.0

    return AcceptanceOut(
        id=acc.id,
        estimate_id=acc.estimate_id,
        contractor_id=acc.contractor_id,
        act_number=acc.act_number,
        period_start=acc.period_start,
        period_end=acc.period_end,
        notes=acc.notes,
        status=acc.status,
        created_at=acc.created_at,
        items_count=items_count,
        total_accepted_value=total_value,
    )


@router.delete("/estimates/{task_id}/acceptances/{acc_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_acceptance(
    task_id: str,
    acc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)
    acc = await _get_acceptance(acc_id, task_id, db)
    await db.delete(acc)
    await db.commit()


# ---------------------------------------------------------------------------
# Work acceptance items
# ---------------------------------------------------------------------------

@router.get("/estimates/{task_id}/acceptances/{acc_id}/contract-items")
async def get_acceptance_contract_items(
    task_id: str,
    acc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Returns estimate items from the acceptance's contractor's contract for this project's estimates."""
    await _get_estimate(task_id, current_user, db)
    acceptance = await _get_acceptance(acc_id, task_id, db)

    # Find the contractor's contracts for this project
    # Get the project_id from task
    from app.models.task import Task as TaskModel
    task = await db.get(TaskModel, task_id)

    if not task or not task.project_id:
        # Fallback: return all estimate items
        result = await db.execute(
            select(EstimateItem).where(
                EstimateItem.task_id == task_id,
                EstimateItem.row_type != "section_header",
            ).order_by(EstimateItem.sort_order, EstimateItem.position)
        )
        items = result.scalars().all()
        return [{"id": i.id, "name": i.name, "unit": i.unit, "quantity": i.quantity, "row_type": i.row_type} for i in items]

    from sqlalchemy import and_

    # Find contracts for this contractor on this project
    contracts_r = await db.execute(
        select(SubcontractorContract).where(
            and_(
                SubcontractorContract.project_id == task.project_id,
                SubcontractorContract.contractor_id == acceptance.contractor_id,
                SubcontractorContract.status == "signed",
            )
        )
    )
    contracts = contracts_r.scalars().all()

    if not contracts:
        # No signed contract - return all items (permissive fallback)
        result = await db.execute(
            select(EstimateItem).where(
                EstimateItem.task_id == task_id,
                EstimateItem.row_type != "section_header",
            ).order_by(EstimateItem.sort_order, EstimateItem.position)
        )
        items = result.scalars().all()
        return [{"id": i.id, "name": i.name, "unit": i.unit, "quantity": i.quantity, "row_type": i.row_type} for i in items]

    # Get estimate_item_ids from contract items (items that have an estimate_item_id)
    contract_ids = [c.id for c in contracts]
    contract_items_r = await db.execute(
        select(SubcontractorContractItem).where(
            and_(
                SubcontractorContractItem.contract_id.in_(contract_ids),
                SubcontractorContractItem.estimate_item_id.isnot(None),
            )
        )
    )
    contract_items = contract_items_r.scalars().all()

    estimate_item_ids = list({ci.estimate_item_id for ci in contract_items})

    if not estimate_item_ids:
        # Contract has no estimate items linked - permissive fallback
        result = await db.execute(
            select(EstimateItem).where(
                EstimateItem.task_id == task_id,
                EstimateItem.row_type != "section_header",
            ).order_by(EstimateItem.sort_order, EstimateItem.position)
        )
        items = result.scalars().all()
        return [{"id": i.id, "name": i.name, "unit": i.unit, "quantity": i.quantity, "row_type": i.row_type} for i in items]

    # Return only the estimate items that are in the contract
    result = await db.execute(
        select(EstimateItem).where(
            and_(
                EstimateItem.id.in_(estimate_item_ids),
                EstimateItem.task_id == task_id,
            )
        ).order_by(EstimateItem.sort_order, EstimateItem.position)
    )
    items = result.scalars().all()

    # Also include contract item quantities for reference
    item_id_to_contract_qty: dict = {}
    for ci in contract_items:
        if ci.estimate_item_id:
            item_id_to_contract_qty[ci.estimate_item_id] = float(ci.quantity)

    return [
        {
            "id": i.id,
            "name": i.name,
            "unit": i.unit,
            "quantity": i.quantity,
            "contract_quantity": item_id_to_contract_qty.get(i.id, i.quantity),
            "row_type": i.row_type,
        }
        for i in items
    ]


@router.get("/estimates/{task_id}/acceptances/{acc_id}/items", response_model=List[AcceptanceItemOut])
async def list_acceptance_items(
    task_id: str,
    acc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)
    await _get_acceptance(acc_id, task_id, db)

    result = await db.execute(
        select(WorkAcceptanceItem)
        .where(WorkAcceptanceItem.acceptance_id == acc_id)
        .order_by(WorkAcceptanceItem.id)
    )
    return result.scalars().all()


@router.put("/estimates/{task_id}/acceptances/{acc_id}/items", response_model=List[AcceptanceItemOut])
async def set_acceptance_items(
    task_id: str,
    acc_id: str,
    items: List[AcceptanceItemIn] = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Replace all items for an acceptance. Any items not present in the payload are deleted."""
    await _get_estimate(task_id, current_user, db)
    await _get_acceptance(acc_id, task_id, db)

    # Validate all estimate_item_ids belong to this estimate
    incoming_item_ids = [i.estimate_item_id for i in items]
    if incoming_item_ids:
        ei_result = await db.execute(
            select(EstimateItem).where(
                EstimateItem.id.in_(incoming_item_ids),
                EstimateItem.task_id == task_id,
            )
        )
        valid_items = {ei.id for ei in ei_result.scalars().all()}
        invalid = set(incoming_item_ids) - valid_items
        if invalid:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Estimate items not found or not part of this estimate: {list(invalid)}",
            )

    # Delete existing items for this acceptance
    existing_result = await db.execute(
        select(WorkAcceptanceItem).where(WorkAcceptanceItem.acceptance_id == acc_id)
    )
    for existing in existing_result.scalars().all():
        await db.delete(existing)

    # Create new items
    new_items: List[WorkAcceptanceItem] = []
    for item_in in items:
        new_item = WorkAcceptanceItem(
            id=str(uuid.uuid4()),
            acceptance_id=acc_id,
            estimate_item_id=item_in.estimate_item_id,
            quantity_accepted=item_in.quantity_accepted,
        )
        db.add(new_item)
        new_items.append(new_item)

    await db.commit()
    for ni in new_items:
        await db.refresh(ni)

    return new_items


@router.delete(
    "/estimates/{task_id}/acceptances/{acc_id}/items/{item_id}",
    status_code=status.HTTP_204_NO_CONTENT,
)
async def delete_acceptance_item(
    task_id: str,
    acc_id: str,
    item_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    await _get_estimate(task_id, current_user, db)
    await _get_acceptance(acc_id, task_id, db)

    result = await db.execute(
        select(WorkAcceptanceItem).where(
            WorkAcceptanceItem.id == item_id,
            WorkAcceptanceItem.acceptance_id == acc_id,
        )
    )
    item = result.scalar_one_or_none()
    if item is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Acceptance item not found")

    await db.delete(item)
    await db.commit()


# ---------------------------------------------------------------------------
# KS-2 export for a work acceptance
# ---------------------------------------------------------------------------

@router.get("/estimates/{task_id}/acceptances/{acc_id}/export-ks2")
async def export_acceptance_ks2(
    task_id: str,
    acc_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate KS-2 Excel document for a work acceptance."""
    from fastapi.responses import StreamingResponse
    import io

    await _get_estimate(task_id, current_user, db)
    acc = await _get_acceptance(acc_id, task_id, db)

    task = await db.get(Task, task_id)
    contractor = await db.get(Contractor, acc.contractor_id) if acc.contractor_id else None

    # Load acceptance items with estimate item data
    items_result = await db.execute(
        select(WorkAcceptanceItem, EstimateItem)
        .join(EstimateItem, WorkAcceptanceItem.estimate_item_id == EstimateItem.id)
        .where(WorkAcceptanceItem.acceptance_id == acc_id)
        .order_by(EstimateItem.position)
    )
    rows = items_result.all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "КС-2"

    thin = Side(style="thin")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value="", bold=False, align="left", border=True):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(name="Arial", size=10, bold=bold)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if border:
            c.border = brd
        return c

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "АКТ о приёмке выполненных работ (форма КС-2)"
    ws["A1"].font = Font(name="Arial", size=13, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Info block
    contractor_name = contractor.name if contractor else "—"
    period = f"{acc.period_start} – {acc.period_end}" if acc.period_start and acc.period_end else "—"
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Подрядчик: {contractor_name}   Акт №: {acc.act_number}   Период: {period}"
    ws["A2"].font = Font(name="Arial", size=10)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Header row
    headers = ["№", "Наименование работ", "Ед.изм.", "Кол-во по дог.", "Выполнено", "Цена за ед.", "Стоимость работ", "Стоимость матер.", "Итого"]
    for col, h in enumerate(headers, start=1):
        cell(3, col, h, bold=True, align="center")
    ws.row_dimensions[3].height = 30

    # Column widths
    widths = [5, 40, 8, 12, 12, 12, 14, 14, 14]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + col)].width = w

    # Data rows
    total_work = 0.0
    total_mat = 0.0
    for idx, (wai, ei) in enumerate(rows, start=1):
        r = 3 + idx
        unit_price = (ei.work_price + ei.mat_price)
        work_val = wai.quantity_accepted * ei.work_price
        mat_val = wai.quantity_accepted * ei.mat_price
        total_val = work_val + mat_val
        total_work += work_val
        total_mat += mat_val

        cell(r, 1, idx, align="center")
        cell(r, 2, ei.name)
        cell(r, 3, ei.unit, align="center")
        cell(r, 4, ei.quantity, align="right")
        cell(r, 5, wai.quantity_accepted, align="right")
        cell(r, 6, round(unit_price, 2), align="right")
        cell(r, 7, round(work_val, 2), align="right")
        cell(r, 8, round(mat_val, 2), align="right")
        cell(r, 9, round(total_val, 2), align="right")
        ws.row_dimensions[r].height = 18

    # Totals row
    tr = 4 + len(rows)
    ws.merge_cells(f"A{tr}:F{tr}")
    c = ws.cell(row=tr, column=1, value="ИТОГО:")
    c.font = Font(name="Arial", size=10, bold=True)
    c.alignment = Alignment(horizontal="right", vertical="center")
    cell(tr, 7, round(total_work, 2), bold=True, align="right")
    cell(tr, 8, round(total_mat, 2), bold=True, align="right")
    cell(tr, 9, round(total_work + total_mat, 2), bold=True, align="right")
    ws.row_dimensions[tr].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"ks2_act_{acc.act_number}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# Acceptance progress
# ---------------------------------------------------------------------------

@router.get("/estimates/{task_id}/acceptance-progress", response_model=List[AcceptanceProgressItem])
async def get_acceptance_progress(
    task_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Returns per-item acceptance progress for all non-header items in the estimate.
    quantity_accepted is the sum across all acceptances (regardless of status).
    """
    await _get_estimate(task_id, current_user, db)

    # Fetch all non-header estimate items for this task
    ei_result = await db.execute(
        select(EstimateItem)
        .where(
            EstimateItem.task_id == task_id,
            EstimateItem.row_type != "section_header",
        )
        .order_by(EstimateItem.position)
    )
    estimate_items = ei_result.scalars().all()

    if not estimate_items:
        return []

    item_ids = [ei.id for ei in estimate_items]

    # Sum accepted quantities per estimate_item_id across all acceptances
    acc_sum_result = await db.execute(
        select(
            WorkAcceptanceItem.estimate_item_id,
            func.coalesce(func.sum(WorkAcceptanceItem.quantity_accepted), 0.0).label("total_accepted"),
        )
        .where(WorkAcceptanceItem.estimate_item_id.in_(item_ids))
        .group_by(WorkAcceptanceItem.estimate_item_id)
    )
    accepted_map: dict[str, float] = {
        row.estimate_item_id: row.total_accepted for row in acc_sum_result.all()
    }

    progress: List[AcceptanceProgressItem] = []
    for ei in estimate_items:
        qty_total = ei.quantity if ei.quantity is not None else 0.0
        qty_accepted = accepted_map.get(ei.id, 0.0)
        qty_remaining = max(qty_total - qty_accepted, 0.0)
        pct_complete = (qty_accepted / qty_total * 100.0) if qty_total > 0 else 0.0

        progress.append(
            AcceptanceProgressItem(
                item_id=ei.id,
                item_name=ei.name,
                unit=ei.unit,
                quantity_total=qty_total,
                quantity_accepted=qty_accepted,
                quantity_remaining=qty_remaining,
                pct_complete=round(pct_complete, 2),
            )
        )

    return progress
