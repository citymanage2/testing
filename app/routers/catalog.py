from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from typing import Optional
from datetime import datetime, timezone
import io
import uuid
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select

import openpyxl
from openpyxl.styles import Font

from app.database import get_db
from app.auth import get_current_user
from app.models.user import User
from app.models.price_catalog import PriceCatalog
from app.models.estimate_item import EstimateItem

router = APIRouter()


class CatalogEntryIn(BaseModel):
    item_type: str = "work"  # work|material
    name: str
    unit: Optional[str] = None
    work_price: float = 0.0
    mat_price: float = 0.0
    tags: Optional[list] = None


class CatalogEntryOut(BaseModel):
    id: str
    item_type: str
    name: str
    unit: Optional[str]
    work_price: float
    mat_price: float
    tags: Optional[list]
    created_at: datetime
    updated_at: datetime


def _out(r: PriceCatalog) -> CatalogEntryOut:
    return CatalogEntryOut(
        id=r.id, item_type=r.item_type, name=r.name, unit=r.unit,
        work_price=r.work_price, mat_price=r.mat_price,
        tags=r.tags or [], created_at=r.created_at,
        updated_at=r.updated_at,
    )


@router.get("", response_model=list[CatalogEntryOut])
async def list_catalog(
    q: Optional[str] = None,
    item_type: Optional[str] = None,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    query = select(PriceCatalog).where(PriceCatalog.user_id == current_user.id)
    if item_type:
        query = query.where(PriceCatalog.item_type == item_type)
    if q:
        query = query.where(PriceCatalog.name.ilike(f"%{q}%"))
    query = query.order_by(PriceCatalog.name).limit(limit)
    rows = (await db.execute(query)).scalars().all()
    return [_out(r) for r in rows]


@router.post("", response_model=CatalogEntryOut)
async def create_catalog_entry(
    body: CatalogEntryIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    row = PriceCatalog(
        id=str(uuid.uuid4()), user_id=current_user.id,
        item_type=body.item_type, name=body.name, unit=body.unit,
        work_price=body.work_price, mat_price=body.mat_price, tags=body.tags or [],
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _out(row)


@router.post("/from-estimate-item/{item_id}", response_model=CatalogEntryOut)
async def save_estimate_item_to_catalog(
    item_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    item = await db.get(EstimateItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    row = PriceCatalog(
        id=str(uuid.uuid4()), user_id=current_user.id,
        item_type="material" if item.type == "Материал" else "work",
        name=item.name, unit=item.unit,
        work_price=item.work_price, mat_price=item.mat_price, tags=[],
    )
    db.add(row)
    await db.commit()
    await db.refresh(row)
    return _out(row)


@router.put("/{entry_id}", response_model=CatalogEntryOut)
async def update_catalog_entry(
    entry_id: str,
    body: CatalogEntryIn,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    row = await db.get(PriceCatalog, entry_id)
    if not row or row.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Not found")
    row.item_type = body.item_type
    row.name = body.name
    row.unit = body.unit
    row.work_price = body.work_price
    row.mat_price = body.mat_price
    row.tags = body.tags or []
    row.updated_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(row)
    return _out(row)


@router.delete("/{entry_id}")
async def delete_catalog_entry(
    entry_id: str,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    row = await db.get(PriceCatalog, entry_id)
    if not row or row.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Not found")
    await db.delete(row)
    await db.commit()
    return {"ok": True}


@router.get("/template")
async def download_catalog_template():
    """Return an Excel template file for catalog import. No auth required."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каталог"

    headers = [
        "Тип (work/material)",
        "Наименование",
        "Единица",
        "Цена работ ₽",
        "Цена материалов ₽",
        "Теги (через запятую)",
    ]
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Example rows
    ws.append(["work", "Укладка плитки", "м²", "1500", "0", "плитка, укладка"])
    ws.append(["material", "Плитка керамическая", "м²", "0", "800", "плитка"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    xlsx_bytes = buffer.read()

    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"catalog_template.xlsx\""},
    )


@router.post("/import-excel")
async def import_catalog_from_excel(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Accept an Excel file upload and upsert catalog entries for the current user."""
    contents = await file.read()
    errors: list[str] = []
    imported = 0
    updated = 0

    try:
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось открыть файл Excel: {exc}")

    ws = wb.active

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Need at least column B (index 1) to be non-empty
        name = row[1] if len(row) > 1 else None
        if not name or not str(name).strip():
            continue

        name = str(name).strip()

        # Column A: item_type
        raw_type = str(row[0]).strip().lower() if row[0] else ""
        item_type = "material" if raw_type == "material" else "work"

        # Column C: unit
        unit = str(row[2]).strip() if len(row) > 2 and row[2] else None

        # Column D: work_price
        try:
            work_price = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
        except (ValueError, TypeError):
            errors.append(f"Строка {row_idx}: неверный формат цены работ «{row[3]}», используется 0.")
            work_price = 0.0

        # Column E: mat_price
        try:
            mat_price = float(row[4]) if len(row) > 4 and row[4] is not None else 0.0
        except (ValueError, TypeError):
            errors.append(f"Строка {row_idx}: неверный формат цены материалов «{row[4]}», используется 0.")
            mat_price = 0.0

        # Column F: tags (comma-separated)
        tags: Optional[list] = None
        if len(row) > 5 and row[5]:
            raw_tags = str(row[5]).strip()
            if raw_tags:
                tags = [t.strip() for t in raw_tags.split(",") if t.strip()]

        # Upsert: find existing entry by user_id + name + item_type
        existing_result = await db.execute(
            select(PriceCatalog).where(
                PriceCatalog.user_id == current_user.id,
                PriceCatalog.name == name,
                PriceCatalog.item_type == item_type,
            )
        )
        existing = existing_result.scalar_one_or_none()

        if existing:
            existing.unit = unit
            existing.work_price = work_price
            existing.mat_price = mat_price
            if tags is not None:
                existing.tags = tags
            existing.updated_at = datetime.now(timezone.utc)
            updated += 1
        else:
            new_entry = PriceCatalog(
                id=str(uuid.uuid4()),
                user_id=current_user.id,
                item_type=item_type,
                name=name,
                unit=unit,
                work_price=work_price,
                mat_price=mat_price,
                tags=tags or [],
            )
            db.add(new_entry)
            imported += 1

    await db.commit()
    return {"imported": imported, "updated": updated, "errors": errors}
