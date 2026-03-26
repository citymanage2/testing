from uuid import UUID
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.auth import get_admin_user, CurrentUser
from app.database import get_db
from app.models.task import Task
from app.models.task_input_file import TaskInputFile
from app.models.price import PriceList
from app.schemas.admin import AdminTaskResponse, AdminTaskDetail, PriceListInfo, TasksPage

router = APIRouter()


@router.get("/tasks", response_model=TasksPage)
async def list_tasks(
    status: str | None = Query(None),
    task_type: str | None = Query(None),
    date_from: datetime | None = Query(None),
    date_to: datetime | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    filters = []
    if status:
        filters.append(Task.status == status)
    if task_type:
        filters.append(Task.task_type == task_type)
    if date_from:
        filters.append(Task.created_at >= date_from)
    if date_to:
        filters.append(Task.created_at <= date_to)

    count_q = select(func.count()).select_from(Task)
    if filters:
        count_q = count_q.where(*filters)
    total_result = await db.execute(count_q)
    total = total_result.scalar_one()

    q = select(Task)
    if filters:
        q = q.where(*filters)
    q = q.order_by(Task.created_at.desc()).offset((page - 1) * page_size).limit(page_size)
    result = await db.execute(q)
    items = result.scalars().all()

    return TasksPage(
        items=[AdminTaskResponse.model_validate(t) for t in items],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/tasks/{task_id}", response_model=AdminTaskDetail)
async def get_task_detail(
    task_id: UUID,
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    return task


@router.delete("/tasks/{task_id}", status_code=204)
async def delete_task(
    task_id: UUID,
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    task = await db.get(Task, task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    await db.delete(task)
    await db.commit()


@router.get("/tasks/{task_id}/download-input/{file_index}")
async def download_input_file(
    task_id: UUID,
    file_index: int,
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(TaskInputFile).where(
            TaskInputFile.task_id == task_id,
            TaskInputFile.file_index == file_index,
        )
    )
    f = result.scalar_one_or_none()
    if not f:
        raise HTTPException(status_code=404, detail="File not found")
    return Response(
        content=f.file_data,
        media_type=f.mime_type,
        headers={"Content-Disposition": f'attachment; filename="{f.file_name}"'},
    )


@router.get("/price-lists/info", response_model=PriceListInfo)
async def get_price_lists_info(
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    async def get_info(type_: str):
        result = await db.execute(
            select(PriceList).where(PriceList.type == type_).order_by(PriceList.updated_at.desc()).limit(1)
        )
        pl = result.scalar_one_or_none()
        if not pl:
            return None
        return {"filename": pl.filename, "updated_at": pl.updated_at.isoformat()}

    return PriceListInfo(
        works=await get_info("works"),
        materials=await get_info("materials"),
    )


@router.post("/price-lists/works")
async def upload_works_price_list(
    file: UploadFile = File(...),
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
    data = await file.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds 20MB limit")

    from app.services.price_service import price_service
    from app.models.price import PriceWork
    import io
    from openpyxl import load_workbook

    # Save raw file
    pl = PriceList(type="works", filename=file.filename, content=data)
    db.add(pl)

    # Parse and upsert price works
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="XLSX file has no active sheet")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid XLSX file: {e}")
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        unit = str(row[1]).strip() if row[1] else ""
        prices_raw = {}
        for i, val in enumerate(row[2:], 1):
            if val is not None:
                try:
                    prices_raw[f"contractor_{i}"] = float(val)
                except (ValueError, TypeError):
                    pass
        min_price = min(prices_raw.values()) if prices_raw else 0.0
        existing = await db.execute(select(PriceWork).where(PriceWork.name == name))
        pw = existing.scalar_one_or_none()
        if pw:
            pw.unit = unit
            pw.prices = prices_raw
            pw.min_price = min_price
        else:
            db.add(PriceWork(name=name, unit=unit, prices=prices_raw, min_price=min_price))

    await db.commit()
    try:
        await price_service.reload_cache()
    except Exception:
        pass  # Data persisted; cache refreshes on next startup
    return {"ok": True}


@router.post("/price-lists/materials")
async def upload_materials_price_list(
    file: UploadFile = File(...),
    admin_user: CurrentUser = Depends(get_admin_user),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files allowed")
    data = await file.read()
    if len(data) > 20 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File exceeds 20MB limit")

    from app.services.price_service import price_service
    from app.models.price import PriceMaterial
    import io
    from openpyxl import load_workbook

    pl = PriceList(type="materials", filename=file.filename, content=data)
    db.add(pl)

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="XLSX file has no active sheet")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid XLSX file: {e}")
    for row in rows:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        unit = str(row[1]).strip() if row[1] else ""
        try:
            price = float(row[2]) if row[2] is not None else 0.0
        except (ValueError, TypeError):
            price = 0.0
        existing = await db.execute(select(PriceMaterial).where(PriceMaterial.name == name))
        pm = existing.scalar_one_or_none()
        if pm:
            pm.unit = unit
            pm.price = price
        else:
            db.add(PriceMaterial(name=name, unit=unit, price=price))

    await db.commit()
    try:
        await price_service.reload_cache()
    except Exception:
        pass  # Data persisted; cache refreshes on next startup
    return {"ok": True}
