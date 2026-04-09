import uuid
from datetime import datetime, timezone
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from pydantic import BaseModel

from app.auth import get_current_user, CurrentUser
from app.database import get_db
from app.models.project import Project
from app.models.work_schedule import WorkScheduleItem, WorkScheduleEntry
from app.models.estimate_item import EstimateItem
from app.models.task import Task

router = APIRouter()


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ScheduleEntryIn(BaseModel):
    period_label: str
    period_type: str
    planned_qty: float = 0.0
    actual_qty: float = 0.0


class ScheduleEntryResponse(BaseModel):
    id: str
    schedule_item_id: str
    period_label: str
    period_type: str
    planned_qty: float
    actual_qty: float

    model_config = {"from_attributes": True}


class ScheduleItemCreate(BaseModel):
    estimate_item_id: Optional[str] = None
    name: str
    unit: Optional[str] = None
    total_quantity: float
    sort_order: Optional[float] = None


class ScheduleItemPatch(BaseModel):
    name: Optional[str] = None
    unit: Optional[str] = None
    total_quantity: Optional[float] = None
    sort_order: Optional[float] = None


class ScheduleItemResponse(BaseModel):
    id: str
    project_id: str
    estimate_item_id: Optional[str]
    name: str
    unit: Optional[str]
    total_quantity: float
    sort_order: float
    entries: list[ScheduleEntryResponse]

    model_config = {"from_attributes": True}


class FromEstimateRequest(BaseModel):
    estimate_ids: list[str]


class SummaryPeriodItem(BaseModel):
    item_id: str
    item_name: str
    planned: float
    actual: float


class SummaryPeriod(BaseModel):
    period_label: str
    total_planned: float
    total_actual: float
    items: list[SummaryPeriodItem]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_project_owned(project_id: str, user_id: str, db: AsyncSession) -> Project:
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project.user_id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    return project


async def _load_item_with_entries(item: WorkScheduleItem, db: AsyncSession) -> ScheduleItemResponse:
    entries_result = await db.execute(
        select(WorkScheduleEntry).where(WorkScheduleEntry.schedule_item_id == item.id)
    )
    entries = entries_result.scalars().all()
    return ScheduleItemResponse(
        id=item.id,
        project_id=item.project_id,
        estimate_item_id=getattr(item, "estimate_item_id", None),
        name=item.name,
        unit=getattr(item, "unit", None),
        total_quantity=item.total_quantity,
        sort_order=getattr(item, "sort_order", 0.0),
        entries=[
            ScheduleEntryResponse(
                id=e.id,
                schedule_item_id=e.schedule_item_id,
                period_label=e.period_label,
                period_type=e.period_type,
                planned_qty=getattr(e, "planned_qty", 0.0),
                actual_qty=getattr(e, "actual_qty", 0.0),
            )
            for e in entries
        ],
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/{project_id}/schedule/items", response_model=list[ScheduleItemResponse])
async def list_schedule_items(
    project_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    result = await db.execute(
        select(WorkScheduleItem)
        .where(WorkScheduleItem.project_id == project_id)
        .order_by(WorkScheduleItem.sort_order)
    )
    items = result.scalars().all()
    return [await _load_item_with_entries(item, db) for item in items]


@router.post("/{project_id}/schedule/items", response_model=ScheduleItemResponse, status_code=201)
async def create_schedule_item(
    project_id: str,
    body: ScheduleItemCreate,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    sort_order = body.sort_order
    if sort_order is None:
        max_result = await db.execute(
            select(func.max(WorkScheduleItem.sort_order)).where(
                WorkScheduleItem.project_id == project_id
            )
        )
        max_val = max_result.scalar()
        sort_order = (max_val or 0.0) + 1.0

    item = WorkScheduleItem(
        id=str(uuid.uuid4()),
        project_id=project_id,
        estimate_item_id=body.estimate_item_id,
        name=body.name,
        unit=body.unit,
        total_quantity=body.total_quantity,
        sort_order=sort_order,
    )
    db.add(item)
    await db.commit()
    await db.refresh(item)
    return await _load_item_with_entries(item, db)


@router.patch("/{project_id}/schedule/items/{item_id}", response_model=ScheduleItemResponse)
async def update_schedule_item(
    project_id: str,
    item_id: str,
    body: ScheduleItemPatch,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    item = await db.get(WorkScheduleItem, item_id)
    if not item or item.project_id != project_id:
        raise HTTPException(status_code=404, detail="Schedule item not found")

    for field, value in body.model_dump(exclude_none=True).items():
        if hasattr(item, field):
            setattr(item, field, value)

    await db.commit()
    await db.refresh(item)
    return await _load_item_with_entries(item, db)


@router.delete("/{project_id}/schedule/items/{item_id}", status_code=204)
async def delete_schedule_item(
    project_id: str,
    item_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    item = await db.get(WorkScheduleItem, item_id)
    if not item or item.project_id != project_id:
        raise HTTPException(status_code=404, detail="Schedule item not found")

    # Cascade delete entries
    entries_result = await db.execute(
        select(WorkScheduleEntry).where(WorkScheduleEntry.schedule_item_id == item_id)
    )
    for entry in entries_result.scalars().all():
        await db.delete(entry)

    await db.delete(item)
    await db.commit()


@router.post("/{project_id}/schedule/items/from-estimates")
async def create_items_from_estimates(
    project_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    """Auto-populate GPR from all signed client estimates in the project.
    Skips items already imported (matched by estimate_item_id) and section headers."""
    await _get_project_owned(project_id, current_user.id, db)

    # Find all signed, non-subcontractor tasks for this project
    tasks_result = await db.execute(
        select(Task).where(
            and_(
                Task.project_id == project_id,
                Task.estimate_status == "signed",
            )
        )
    )
    tasks = [t for t in tasks_result.scalars().all() if getattr(t, "estimate_type", None) != "subcontractor"]

    if not tasks:
        # Fall back to all completed tasks (draft) if none signed
        tasks_result2 = await db.execute(
            select(Task).where(Task.project_id == project_id)
        )
        tasks = [t for t in tasks_result2.scalars().all() if getattr(t, "estimate_type", None) != "subcontractor"]

    task_ids = [t.id for t in tasks]
    if not task_ids:
        return {"created": 0}

    # Get already-imported estimate_item_ids to avoid duplicates
    existing_result = await db.execute(
        select(WorkScheduleItem.estimate_item_id).where(
            and_(
                WorkScheduleItem.project_id == project_id,
                WorkScheduleItem.estimate_item_id.isnot(None),
            )
        )
    )
    already_imported = {row[0] for row in existing_result.all()}

    # Get max sort_order
    max_result = await db.execute(
        select(func.max(WorkScheduleItem.sort_order)).where(
            WorkScheduleItem.project_id == project_id
        )
    )
    next_sort = (max_result.scalar() or 0.0) + 1.0

    # Load all non-header work items from those tasks
    estimate_items_result = await db.execute(
        select(EstimateItem).where(
            and_(
                EstimateItem.task_id.in_(task_ids),
                EstimateItem.type == "Работа",
            )
        ).order_by(EstimateItem.sort_order)
    )
    estimate_items = [
        i for i in estimate_items_result.scalars().all()
        if getattr(i, "row_type", "item") != "section_header"
        and i.id not in already_imported
    ]

    count = 0
    for est_item in estimate_items:
        new_item = WorkScheduleItem(
            id=str(uuid.uuid4()),
            project_id=project_id,
            estimate_item_id=est_item.id,
            name=est_item.name,
            unit=est_item.unit or "",
            total_quantity=float(est_item.quantity or 0),
            sort_order=next_sort,
        )
        db.add(new_item)
        next_sort += 1.0
        count += 1

    await db.commit()
    return {"created": count}


@router.post("/{project_id}/schedule/items/from-estimate")
async def create_items_from_estimate(
    project_id: str,
    body: FromEstimateRequest,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    # Get existing sort_order max
    max_result = await db.execute(
        select(func.max(WorkScheduleItem.sort_order)).where(
            WorkScheduleItem.project_id == project_id
        )
    )
    next_sort = (max_result.scalar() or 0.0) + 1.0

    estimate_items_result = await db.execute(
        select(EstimateItem).where(
            and_(
                EstimateItem.task_id.in_(body.estimate_ids),
                EstimateItem.type == "Работа",
            )
        )
    )
    estimate_items = estimate_items_result.scalars().all()

    count = 0
    for est_item in estimate_items:
        new_item = WorkScheduleItem(
            id=str(uuid.uuid4()),
            project_id=project_id,
            estimate_item_id=est_item.id,
            name=est_item.name,
            unit=est_item.unit or "",
            total_quantity=est_item.quantity,
            sort_order=next_sort,
        )
        db.add(new_item)
        next_sort += 1.0
        count += 1

    await db.commit()
    return {"created": count}


@router.put(
    "/{project_id}/schedule/items/{item_id}/entries",
    response_model=list[ScheduleEntryResponse],
)
async def replace_entries(
    project_id: str,
    item_id: str,
    body: list[ScheduleEntryIn],
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    item = await db.get(WorkScheduleItem, item_id)
    if not item or item.project_id != project_id:
        raise HTTPException(status_code=404, detail="Schedule item not found")

    # Delete existing entries
    existing_result = await db.execute(
        select(WorkScheduleEntry).where(WorkScheduleEntry.schedule_item_id == item_id)
    )
    for entry in existing_result.scalars().all():
        await db.delete(entry)

    # Insert new entries
    new_entries = []
    for entry_in in body:
        new_entry = WorkScheduleEntry(
            id=str(uuid.uuid4()),
            schedule_item_id=item_id,
            period_label=entry_in.period_label,
            period_type=entry_in.period_type,
            planned_qty=entry_in.planned_qty,
            actual_qty=entry_in.actual_qty,
        )
        db.add(new_entry)
        new_entries.append(new_entry)

    await db.commit()

    return [
        ScheduleEntryResponse(
            id=e.id,
            schedule_item_id=e.schedule_item_id,
            period_label=e.period_label,
            period_type=e.period_type,
            planned_qty=e.planned_qty,
            actual_qty=e.actual_qty,
        )
        for e in new_entries
    ]


@router.get("/{project_id}/schedule/summary", response_model=list[SummaryPeriod])
async def schedule_summary(
    project_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    items_result = await db.execute(
        select(WorkScheduleItem).where(WorkScheduleItem.project_id == project_id)
    )
    items = items_result.scalars().all()

    if not items:
        return []

    item_ids = [i.id for i in items]
    item_map = {i.id: i for i in items}

    entries_result = await db.execute(
        select(WorkScheduleEntry).where(WorkScheduleEntry.schedule_item_id.in_(item_ids))
    )
    entries = entries_result.scalars().all()

    # Group by period_label
    periods: dict[str, dict] = {}
    for entry in entries:
        label = entry.period_label
        if label not in periods:
            periods[label] = {"total_planned": 0.0, "total_actual": 0.0, "items": {}}

        p = periods[label]
        p["total_planned"] += getattr(entry, "planned_qty", 0.0)
        p["total_actual"] += getattr(entry, "actual_qty", 0.0)

        sched_item = item_map.get(entry.schedule_item_id)
        if sched_item:
            iid = sched_item.id
            if iid not in p["items"]:
                p["items"][iid] = {"item_id": iid, "item_name": sched_item.name, "planned": 0.0, "actual": 0.0}
            p["items"][iid]["planned"] += getattr(entry, "planned_qty", 0.0)
            p["items"][iid]["actual"] += getattr(entry, "actual_qty", 0.0)

    result = []
    for period_label, data in sorted(periods.items()):
        result.append(
            SummaryPeriod(
                period_label=period_label,
                total_planned=data["total_planned"],
                total_actual=data["total_actual"],
                items=[
                    SummaryPeriodItem(**item_data)
                    for item_data in data["items"].values()
                ],
            )
        )
    return result


@router.get("/{project_id}/schedule/export-excel")
async def export_schedule_excel(
    project_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    """Export GPR (work schedule) to Excel."""
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    # verify project ownership
    project = await db.get(Project, project_id)
    if not project or project.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Project not found")

    # Load schedule items with entries
    items_result = await db.execute(
        select(WorkScheduleItem)
        .where(WorkScheduleItem.project_id == project_id)
        .order_by(WorkScheduleItem.sort_order)
    )
    items = items_result.scalars().all()

    # Get all unique periods across all items
    all_periods = set()
    item_entries = {}
    for item in items:
        entries_result = await db.execute(
            select(WorkScheduleEntry).where(WorkScheduleEntry.schedule_item_id == item.id)
        )
        entries = entries_result.scalars().all()
        item_entries[item.id] = {e.period_label: e for e in entries}
        for e in entries:
            all_periods.add(e.period_label)
    periods = sorted(all_periods)

    wb = Workbook()
    ws = wb.active
    ws.title = "ГПР"

    thin = Side(style="thin")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1565C0")

    # Title
    ws.merge_cells(f"A1:{chr(65 + 3 + len(periods))}1")
    t = ws["A1"]
    t.value = f"График производства работ — {project.name}"
    t.font = Font(name="Arial", size=13, bold=True)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Headers
    headers = ["№", "Наименование", "Ед.", "Кол-во"] + periods
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
    ws.row_dimensions[2].height = 32
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    for i, _ in enumerate(periods):
        ws.column_dimensions[chr(69 + i)].width = 10

    # Data rows
    for idx, item in enumerate(items, 1):
        r = 2 + idx
        ws.cell(row=r, column=1, value=idx).border = brd
        ws.cell(row=r, column=2, value=item.name).border = brd
        ws.cell(row=r, column=3, value=item.unit or "").border = brd
        ws.cell(row=r, column=4, value=float(item.total_quantity or 0)).border = brd
        for pi, period in enumerate(periods):
            entry = item_entries[item.id].get(period)
            planned = float(entry.planned_qty or 0) if entry else 0
            actual = float(entry.actual_qty or 0) if entry else 0
            cell_val = f"П: {planned}" + (f"\nФ: {actual}" if actual else "")
            c = ws.cell(row=r, column=5 + pi, value=cell_val)
            c.border = brd
            c.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="gpr_{project_id}.xlsx"'},
    )
