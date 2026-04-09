import io
import uuid
from datetime import datetime, timezone, date
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from pydantic import BaseModel

from app.auth import get_current_user, CurrentUser
from app.database import get_db
from app.models.project import Project
from app.models.client_act import ClientKs2Act, ClientKs2ActItem
from app.models.estimate_item import EstimateItem
from app.models.task import Task

router = APIRouter()


# ---------------------------------------------------------------------------
# Schemas
# ---------------------------------------------------------------------------

class ActCreate(BaseModel):
    act_number: str
    period_start: Optional[date] = None
    period_end: Optional[date] = None
    contractor_id: Optional[str] = None
    notes: Optional[str] = None


class ActPatch(BaseModel):
    act_number: Optional[str] = None
    status: Optional[str] = None
    notes: Optional[str] = None
    signed_at: Optional[datetime] = None
    contractor_id: Optional[str] = None
    period_start: Optional[date] = None
    period_end: Optional[date] = None


class ActResponse(BaseModel):
    id: str
    project_id: str
    act_number: str
    status: str
    period_start: Optional[date]
    period_end: Optional[date]
    contractor_id: Optional[str]
    notes: Optional[str]
    signed_at: Optional[datetime]
    items_count: int
    total_amount: float
    created_at: Optional[datetime]

    model_config = {"from_attributes": True}


class ActItemIn(BaseModel):
    estimate_item_id: str
    quantity_presented: float
    unit_price: Optional[float] = None


class ActItemResponse(BaseModel):
    id: str
    act_id: str
    estimate_item_id: str
    estimate_item_name: Optional[str]
    estimate_item_unit: Optional[str]
    estimate_item_quantity_total: Optional[float]
    already_actioned_qty: float
    remaining_qty: float
    quantity_presented: float
    unit_price: float
    total: float

    model_config = {"from_attributes": True}


class ActioningSummaryItem(BaseModel):
    estimate_item_id: str
    name: str
    unit: str
    quantity_total: float
    quantity_actioned: float
    quantity_remaining: float
    pct_actioned: float


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

async def _get_project_owned(project_id: str, user_id: str, db: AsyncSession) -> Project:
    project = await db.get(Project, project_id)
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if project.user_id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    return project


async def _enrich_act(act: ClientKs2Act, db: AsyncSession) -> ActResponse:
    items_result = await db.execute(
        select(ClientKs2ActItem).where(ClientKs2ActItem.act_id == act.id)
    )
    items = items_result.scalars().all()
    items_count = len(items)
    total_amount = sum(
        getattr(i, "quantity_presented", 0.0) * getattr(i, "unit_price", 0.0)
        for i in items
    )
    return ActResponse(
        id=act.id,
        project_id=act.project_id,
        act_number=act.act_number,
        status=act.status,
        period_start=getattr(act, "period_start", None),
        period_end=getattr(act, "period_end", None),
        contractor_id=getattr(act, "contractor_id", None),
        notes=getattr(act, "notes", None),
        signed_at=getattr(act, "signed_at", None),
        items_count=items_count,
        total_amount=total_amount,
        created_at=getattr(act, "created_at", None),
    )


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/{project_id}/client-acts", response_model=list[ActResponse])
async def list_acts(
    project_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    result = await db.execute(
        select(ClientKs2Act).where(ClientKs2Act.project_id == project_id)
    )
    acts = result.scalars().all()
    return [await _enrich_act(act, db) for act in acts]


@router.post("/{project_id}/client-acts", response_model=ActResponse, status_code=201)
async def create_act(
    project_id: str,
    body: ActCreate,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    act = ClientKs2Act(
        id=str(uuid.uuid4()),
        project_id=project_id,
        act_number=body.act_number,
        status="draft",
        period_start=body.period_start,
        period_end=body.period_end,
        contractor_id=body.contractor_id,
        notes=body.notes,
        created_at=datetime.now(timezone.utc),
    )
    db.add(act)
    await db.commit()
    await db.refresh(act)
    return await _enrich_act(act, db)


@router.patch("/{project_id}/client-acts/{act_id}", response_model=ActResponse)
async def update_act(
    project_id: str,
    act_id: str,
    body: ActPatch,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    act = await db.get(ClientKs2Act, act_id)
    if not act or act.project_id != project_id:
        raise HTTPException(status_code=404, detail="Act not found")

    updates = body.model_dump(exclude_none=True)

    if "status" in updates and updates["status"] == "signed":
        if not updates.get("signed_at") and not getattr(act, "signed_at", None):
            act.signed_at = datetime.now(timezone.utc)

    for field, value in updates.items():
        if hasattr(act, field):
            setattr(act, field, value)

    await db.commit()
    await db.refresh(act)
    return await _enrich_act(act, db)


@router.delete("/{project_id}/client-acts/{act_id}", status_code=204)
async def delete_act(
    project_id: str,
    act_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    act = await db.get(ClientKs2Act, act_id)
    if not act or act.project_id != project_id:
        raise HTTPException(status_code=404, detail="Act not found")

    if act.status != "draft":
        raise HTTPException(status_code=400, detail="Only draft acts can be deleted")

    items_result = await db.execute(
        select(ClientKs2ActItem).where(ClientKs2ActItem.act_id == act_id)
    )
    for item in items_result.scalars().all():
        await db.delete(item)

    await db.delete(act)
    await db.commit()


@router.get("/{project_id}/client-acts/{act_id}/items", response_model=list[ActItemResponse])
async def list_act_items(
    project_id: str,
    act_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    act = await db.get(ClientKs2Act, act_id)
    if not act or act.project_id != project_id:
        raise HTTPException(status_code=404, detail="Act not found")

    items_result = await db.execute(
        select(ClientKs2ActItem).where(ClientKs2ActItem.act_id == act_id)
    )
    items = items_result.scalars().all()

    response_items = []
    for item in items:
        est_item = await db.get(EstimateItem, item.estimate_item_id)

        # Sum from other non-cancelled acts
        already_result = await db.execute(
            select(func.sum(ClientKs2ActItem.quantity_presented)).where(
                and_(
                    ClientKs2ActItem.estimate_item_id == item.estimate_item_id,
                    ClientKs2ActItem.act_id != act_id,
                )
            ).join(
                ClientKs2Act, ClientKs2Act.id == ClientKs2ActItem.act_id
            ).where(
                and_(
                    ClientKs2Act.project_id == project_id,
                    ClientKs2Act.status != "cancelled",
                )
            )
        )
        already_actioned = already_result.scalar() or 0.0
        est_qty = est_item.quantity if est_item else 0.0
        remaining = est_qty - already_actioned

        response_items.append(
            ActItemResponse(
                id=item.id,
                act_id=item.act_id,
                estimate_item_id=item.estimate_item_id,
                estimate_item_name=est_item.name if est_item else None,
                estimate_item_unit=est_item.unit if est_item else None,
                estimate_item_quantity_total=est_qty if est_item else None,
                already_actioned_qty=already_actioned,
                remaining_qty=remaining,
                quantity_presented=item.quantity_presented,
                unit_price=item.unit_price,
                total=item.quantity_presented * item.unit_price,
            )
        )

    return response_items


@router.put("/{project_id}/client-acts/{act_id}/items", response_model=list[ActItemResponse])
async def replace_act_items(
    project_id: str,
    act_id: str,
    body: list[ActItemIn],
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    act = await db.get(ClientKs2Act, act_id)
    if not act or act.project_id != project_id:
        raise HTTPException(status_code=404, detail="Act not found")

    # Validate each line
    for line in body:
        est_item = await db.get(EstimateItem, line.estimate_item_id)
        if not est_item:
            raise HTTPException(
                status_code=400,
                detail=f"Estimate item {line.estimate_item_id} not found",
            )
        # Ensure the item belongs to a CLIENT (not subcontractor) estimate
        source_task = await db.get(Task, est_item.task_id)
        if source_task and getattr(source_task, "estimate_type", None) == "subcontractor":
            raise HTTPException(
                status_code=400,
                detail=f"Позиция '{est_item.name}' принадлежит смете субподрядчика. "
                       "КС-2 с заказчиком заполняется только по клиентской смете.",
            )

        # Sum from other non-cancelled acts (exclude current act)
        already_result = await db.execute(
            select(func.sum(ClientKs2ActItem.quantity_presented)).select_from(
                ClientKs2ActItem
            ).join(
                ClientKs2Act, ClientKs2Act.id == ClientKs2ActItem.act_id
            ).where(
                and_(
                    ClientKs2ActItem.estimate_item_id == line.estimate_item_id,
                    ClientKs2ActItem.act_id != act_id,
                    ClientKs2Act.project_id == project_id,
                    ClientKs2Act.status != "cancelled",
                )
            )
        )
        already_actioned = already_result.scalar() or 0.0
        remaining = est_item.quantity - already_actioned

        if line.quantity_presented > remaining:
            raise HTTPException(
                status_code=400,
                detail=f"Превышен доступный остаток по позиции '{est_item.name}': "
                       f"доступно {remaining}",
            )

    # Delete existing items
    existing_result = await db.execute(
        select(ClientKs2ActItem).where(ClientKs2ActItem.act_id == act_id)
    )
    for old_item in existing_result.scalars().all():
        await db.delete(old_item)

    # Insert new items
    new_items = []
    for line in body:
        est_item = await db.get(EstimateItem, line.estimate_item_id)
        unit_price = line.unit_price
        if unit_price is None:
            # Fall back to sale_price on EstimateItem or work_price
            unit_price = getattr(est_item, "sale_price", None) or (
                est_item.work_price + est_item.mat_price if est_item else 0.0
            )

        new_item = ClientKs2ActItem(
            id=str(uuid.uuid4()),
            act_id=act_id,
            estimate_item_id=line.estimate_item_id,
            quantity_presented=line.quantity_presented,
            unit_price=unit_price,
        )
        db.add(new_item)
        new_items.append((new_item, est_item))

    await db.commit()

    response_items = []
    for item, est_item in new_items:
        already_result = await db.execute(
            select(func.sum(ClientKs2ActItem.quantity_presented)).select_from(
                ClientKs2ActItem
            ).join(
                ClientKs2Act, ClientKs2Act.id == ClientKs2ActItem.act_id
            ).where(
                and_(
                    ClientKs2ActItem.estimate_item_id == item.estimate_item_id,
                    ClientKs2ActItem.act_id != act_id,
                    ClientKs2Act.project_id == project_id,
                    ClientKs2Act.status != "cancelled",
                )
            )
        )
        already_actioned = already_result.scalar() or 0.0
        est_qty = est_item.quantity if est_item else 0.0

        response_items.append(
            ActItemResponse(
                id=item.id,
                act_id=item.act_id,
                estimate_item_id=item.estimate_item_id,
                estimate_item_name=est_item.name if est_item else None,
                estimate_item_unit=est_item.unit if est_item else None,
                estimate_item_quantity_total=est_qty if est_item else None,
                already_actioned_qty=already_actioned,
                remaining_qty=est_qty - already_actioned,
                quantity_presented=item.quantity_presented,
                unit_price=item.unit_price,
                total=item.quantity_presented * item.unit_price,
            )
        )

    return response_items


@router.get(
    "/{project_id}/actioning-summary",
    response_model=list[ActioningSummaryItem],
)
async def actioning_summary(
    project_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)

    # Get only CLIENT (non-subcontractor) tasks for this project
    tasks_result = await db.execute(
        select(Task).where(Task.project_id == project_id)
    )
    client_tasks = [
        t for t in tasks_result.scalars().all()
        if getattr(t, "estimate_type", None) != "subcontractor"
    ]
    task_ids = [t.id for t in client_tasks]

    if not task_ids:
        return []

    # Get all non-header estimate items for this project
    items_result = await db.execute(
        select(EstimateItem).where(
            and_(
                EstimateItem.task_id.in_(task_ids),
                EstimateItem.row_type != "section_header",
            )
        ).order_by(EstimateItem.sort_order, EstimateItem.position)
    )
    estimate_items = items_result.scalars().all()

    # Get all actioned quantities from non-cancelled acts
    actioned_result = await db.execute(
        select(
            ClientKs2ActItem.estimate_item_id,
            func.sum(ClientKs2ActItem.quantity_presented).label("total_actioned"),
        ).select_from(ClientKs2ActItem).join(
            ClientKs2Act, ClientKs2Act.id == ClientKs2ActItem.act_id
        ).where(
            and_(
                ClientKs2Act.project_id == project_id,
                ClientKs2Act.status != "cancelled",
            )
        ).group_by(ClientKs2ActItem.estimate_item_id)
    )
    actioned_map: dict[str, float] = {
        row.estimate_item_id: float(row.total_actioned)
        for row in actioned_result.fetchall()
    }

    summary = []
    for ei in estimate_items:
        actioned = actioned_map.get(ei.id, 0.0)
        remaining = ei.quantity - actioned
        pct = (actioned / ei.quantity * 100.0) if ei.quantity else 0.0
        summary.append(
            ActioningSummaryItem(
                estimate_item_id=ei.id,
                name=ei.name,
                unit=ei.unit or "",
                quantity_total=ei.quantity,
                quantity_actioned=actioned,
                quantity_remaining=remaining,
                pct_actioned=round(pct, 2),
            )
        )

    return summary


@router.get("/{project_id}/client-acts/{act_id}/export-ks2")
async def export_ks2(
    project_id: str,
    act_id: str,
    current_user: CurrentUser,
    db: AsyncSession = Depends(get_db),
):
    await _get_project_owned(project_id, current_user.id, db)
    act = await db.get(ClientKs2Act, act_id)
    if not act or act.project_id != project_id:
        raise HTTPException(status_code=404, detail="Act not found")

    project = await db.get(Project, project_id)

    # Load act items with estimate item data
    items_result = await db.execute(
        select(ClientKs2ActItem).where(ClientKs2ActItem.act_id == act_id)
    )
    act_items = items_result.scalars().all()

    # Build Excel KS-2 form
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "КС-2"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def bc(row, col, value, bold=False, size=10, wrap=True, align='left', merge_to=None, fill=None, border_style=True):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name='Arial', size=size, bold=bold)
        cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
        if border_style:
            cell.border = border
        if fill:
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
        if merge_to:
            ws.merge_cells(start_row=row, start_column=col, end_row=merge_to[0], end_column=merge_to[1])
        return cell

    # --- Title block ---
    ws.merge_cells('A1:I1')
    ws['A1'] = 'АКТ\nо приёмке выполненных работ'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 40

    ws.merge_cells('A2:I2')
    ws['A2'] = 'Форма № КС-2'
    ws['A2'].font = Font(name='Arial', size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='right', vertical='center')

    # Customer/contractor info
    project_name = getattr(project, 'name', '') if project else ''

    ws.merge_cells('A3:C3')
    ws['A3'] = 'Заказчик:'
    ws['A3'].font = Font(name='Arial', size=10, bold=True)
    ws.merge_cells('D3:I3')
    ws['D3'] = '_______________________________________'

    ws.merge_cells('A4:C4')
    ws['A4'] = 'Подрядчик:'
    ws['A4'].font = Font(name='Arial', size=10, bold=True)
    ws.merge_cells('D4:I4')
    ws['D4'] = '_______________________________________'

    ws.merge_cells('A5:C5')
    ws['A5'] = 'Объект:'
    ws['A5'].font = Font(name='Arial', size=10, bold=True)
    ws.merge_cells('D5:I5')
    ws['D5'] = project_name

    # Contract & period info
    contract_num = getattr(project, 'contract_number', '') if project else ''
    contract_date = getattr(project, 'contract_date', '') if project else ''
    ws.merge_cells('A6:I6')
    ws['A6'] = f'Договор подряда № {contract_num or "___"} от {contract_date or "___"}'
    ws['A6'].font = Font(name='Arial', size=10)

    # Act info row
    ws['A7'] = '№ документа'
    ws['A7'].font = Font(name='Arial', size=9, bold=True)
    ws['A7'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws['A7'].border = border
    ws.merge_cells('B7:C7')
    ws['B7'] = 'Дата составления'
    ws['B7'].font = Font(name='Arial', size=9, bold=True)
    ws['B7'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws['B7'].border = border
    ws.merge_cells('D7:F7')
    ws['D7'] = 'Отчётный период'
    ws['D7'].font = Font(name='Arial', size=9, bold=True)
    ws['D7'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws['D7'].border = border
    ws.merge_cells('G7:I7')
    ws['G7'] = 'Сметная (договорная) стоимость'
    ws['G7'].font = Font(name='Arial', size=9, bold=True)
    ws['G7'].alignment = Alignment(horizontal='center', wrap_text=True)
    ws['G7'].border = border

    ws['A8'] = act.act_number or ''
    ws['A8'].alignment = Alignment(horizontal='center')
    ws['A8'].border = border
    ws.merge_cells('B8:C8')
    ws['B8'] = str(act.signed_at or '')
    ws['B8'].alignment = Alignment(horizontal='center')
    ws['B8'].border = border
    ws.merge_cells('D8:F8')
    period = f"{act.period_start} — {act.period_end}" if act.period_start else ''
    ws['D8'] = period
    ws['D8'].alignment = Alignment(horizontal='center')
    ws['D8'].border = border
    ws.merge_cells('G8:I8')
    ws['G8'] = ''
    ws['G8'].border = border

    # Table headers row 9
    headers = [
        ('A', 9, '№\nп/п', 4),
        ('B', 9, 'Шифр и №\nпозиции', 12),
        ('C', 9, 'Наименование работ и затрат', 40),
        ('D', 9, 'Ед.\nизм.', 8),
        ('E', 9, 'Кол-во\n(всего)', 10),
        ('F', 9, 'Цена\nед.', 12),
        ('G', 9, 'Сумма\n(всего)', 14),
        ('H', 9, 'Кол-во\n(период)', 10),
        ('I', 9, 'Стоимость\n(период)', 14),
    ]
    for i, (col, _, label, width) in enumerate(headers):
        ws.column_dimensions[col].width = width
        c = ws.cell(row=9, column=i + 1, value=label)
        c.font = Font(name='Arial', size=9, bold=True)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        c.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    ws.row_dimensions[9].height = 36

    # Data rows starting from row 10
    data_row = 10
    total_all = 0.0
    total_period = 0.0

    for idx, item in enumerate(act_items, 1):
        est = await db.get(EstimateItem, item.estimate_item_id) if item.estimate_item_id else None
        name = est.name if est else ''
        unit = est.unit if est else ''
        total_qty = float(est.quantity) if est and est.quantity else 0.0
        unit_price = float(item.unit_price) if item.unit_price else 0.0
        qty_period = float(item.quantity_presented)
        sum_all = total_qty * unit_price
        sum_period = qty_period * unit_price
        total_all += sum_all
        total_period += sum_period

        row_data = [idx, '', name, unit, total_qty, unit_price, round(sum_all, 2), qty_period, round(sum_period, 2)]
        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=data_row, column=col_idx, value=val)
            c.font = Font(name='Arial', size=9)
            c.border = border
            if col_idx in (1, 2, 4, 5, 6, 7, 8, 9):
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if col_idx in (6, 7, 9):
                c.number_format = '#,##0.00'
        ws.row_dimensions[data_row].height = 30
        data_row += 1

    # Totals
    vat_rate = 0.20
    vat_all = round(total_all * vat_rate, 2)
    vat_period = round(total_period * vat_rate, 2)
    total_all_vat = round(total_all + vat_all, 2)
    total_period_vat = round(total_period + vat_period, 2)

    def total_row(row, label, g_val, i_val, bold=False):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name='Arial', size=9, bold=bold)
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = border
        for col in range(2, 7):
            ws.cell(row=row, column=col).border = border
        for col_idx, val in [(7, g_val), (8, ''), (9, i_val)]:
            c2 = ws.cell(row=row, column=col_idx, value=val)
            c2.font = Font(name='Arial', size=9, bold=bold)
            c2.alignment = Alignment(horizontal='center', vertical='center')
            c2.border = border
            c2.number_format = '#,##0.00'

    total_row(data_row, 'ИТОГО:', round(total_all, 2), round(total_period, 2), bold=True)
    total_row(data_row + 1, f'в том числе НДС {int(vat_rate * 100)}%:', vat_all, vat_period)
    total_row(data_row + 2, 'ИТОГО с НДС:', total_all_vat, total_period_vat, bold=True)

    # Signatures
    sig_row = data_row + 4
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=4)
    ws.cell(row=sig_row, column=1, value='Сдал: _____________ / _____________ / _____________').font = Font(name='Arial', size=9)
    ws.merge_cells(start_row=sig_row, start_column=6, end_row=sig_row, end_column=9)
    ws.cell(row=sig_row, column=6, value='Принял: _____________ / _____________ / _____________').font = Font(name='Arial', size=9)
    ws.row_dimensions[sig_row].height = 20

    # Output
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="KS2_{act.act_number or act_id[:8]}.xlsx"'},
    )
