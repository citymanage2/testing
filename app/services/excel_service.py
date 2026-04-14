"""Generate and parse Excel estimate/VOR files."""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL   = PatternFill(fill_type="solid", fgColor="D9E1F2")
_OPT_FILL      = PatternFill(fill_type="solid", fgColor="FFF3CD")
_ANALOGUE_FILL = PatternFill(fill_type="solid", fgColor="C8E6C9")
_ESTIMATED_FILL = PatternFill(fill_type="solid", fgColor="FFE0B2")   # orange — AI-estimated price
_SCAN_ERR_FILL  = PatternFill(fill_type="solid", fgColor="FFCDD2")   # red — scan arithmetic error
_DISCREP_FILL   = PatternFill(fill_type="solid", fgColor="FFF9C4")   # yellow — cross-source discrepancy
_BOLD = Font(bold=True)


# ---------------------------------------------------------------------------
# VOR Excel  (Модули 2.1 / 2.2 / 2.3)
# ---------------------------------------------------------------------------

def build_vor_excel(items, task_type: str = "") -> bytes:
    """
    Build vor.xlsx for ВОР task types.

    Sheet 1: «ВОР» — all positions with quantities.
    Sheet 2: «Несоответствия» — positions with discrepancy=True (modules 2.2).

    Arithmetic (totals, counts) is computed here, not by Claude.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ВОР"

    headers = [
        "№", "Раздел", "Тип", "Наименование", "Ед.", "Кол-во (ТЗ)", "Кол-во (Проект)",
        "Кол-во итог", "Примечание",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    discrepancy_items = []
    for i, item in enumerate(items, 1):
        qty_tz      = getattr(item, "qty_from_tz", None)
        qty_proj    = getattr(item, "qty_from_project", None)
        qty_final   = item.quantity
        has_discrep = getattr(item, "discrepancy", False)

        row_data = [
            i,
            item.section,
            item.type,
            item.name,
            item.unit,
            qty_tz if qty_tz is not None else "",
            qty_proj if qty_proj is not None else "",
            qty_final,
            item.comment or "",
        ]
        ws.append(row_data)
        row_num = ws.max_row

        if has_discrep:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = _DISCREP_FILL
            discrepancy_items.append(item)

    # Footer: total count and items without qty
    ws.append([])
    total_row = ws.max_row + 1
    total_items = len(items)
    without_qty = sum(1 for it in items if it.quantity is None or it.quantity == 0)
    ws.cell(row=total_row, column=4, value="Всего позиций:").font = _BOLD
    ws.cell(row=total_row, column=8, value=total_items).font = _BOLD
    if without_qty:
        ws.cell(row=total_row + 1, column=4, value="Без объёма:").font = Font(bold=True, color="C0392B")
        ws.cell(row=total_row + 1, column=8, value=without_qty).font = Font(bold=True, color="C0392B")

    _set_vor_col_widths(ws)

    # Sheet 2: discrepancies (module 2.2)
    if discrepancy_items:
        wd = wb.create_sheet("Несоответствия")
        d_headers = ["№", "Наименование", "Ед.", "Кол-во ТЗ", "Кол-во Проект", "Разница", "Примечание"]
        for col, h in enumerate(d_headers, 1):
            cell = wd.cell(row=1, column=col, value=h)
            cell.font = _BOLD
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for i, item in enumerate(discrepancy_items, 1):
            qty_tz   = getattr(item, "qty_from_tz", None) or 0
            qty_proj = getattr(item, "qty_from_project", None) or 0
            diff     = qty_tz - qty_proj
            wd.append([i, item.name, item.unit, qty_tz, qty_proj, diff, item.comment or ""])
        for col, w in enumerate([5, 50, 8, 14, 14, 14, 40], 1):
            wd.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _set_vor_col_widths(ws) -> None:
    for col, w in enumerate([5, 20, 12, 50, 8, 14, 14, 14, 40], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Estimate Excel  (Модули 2.4 – 2.8)
# ---------------------------------------------------------------------------

def build_estimate_excel(items, filter_type: str = "all") -> bytes:
    """
    Build smeta.xlsx for estimate task types.

    Sheet 1: «Смета» — all positions with live formulas for totals.
    Sheet 2: «Требуют проверки» — positions with is_estimated=True.

    Rules:
    - Totals in Excel are openpyxl formulas (=qty*price), not hardcoded numbers.
    - is_estimated rows are highlighted orange.
    - scan_math_error rows are highlighted red.
    - A separate sheet lists all AI-estimated positions for human review.
    """
    if filter_type == "works":
        items = [i for i in items if i.type in ("Работа", "work")]
    elif filter_type == "materials":
        items = [i for i in items if i.type in ("Материал", "material")]

    wb = Workbook()
    ws = wb.active
    ws.title = {"works": "Работы", "materials": "Материалы"}.get(filter_type, "Смета")

    headers = [
        "№", "Раздел", "Тип", "Наименование", "Ед.", "Кол-во",
        "Цена работ", "Цена мат.", "Стоимость", "Источник", "Флаги",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    estimated_items = []
    data_start_row = 2

    for item in items:
        row_num = ws.max_row + 1

        is_estimated   = getattr(item, "is_estimated", False)
        scan_err       = getattr(item, "scan_math_error", False)
        source_display = getattr(item, "source", None) or getattr(item, "source_url", None) or ""

        flags = []
        if is_estimated:
            flags.append("AI-оценка")
        if scan_err:
            flags.append("ошибка скана")
        if getattr(item, "discrepancy", False):
            flags.append("расхождение")

        # Live formula: qty * (work_price + mat_price)
        qty_col    = get_column_letter(6)   # F
        wp_col     = get_column_letter(7)   # G
        mp_col     = get_column_letter(8)   # H
        total_formula = f"={qty_col}{row_num}*({wp_col}{row_num}+{mp_col}{row_num})"

        ws.append([
            item.position,
            item.section,
            item.type,
            item.name,
            item.unit,
            item.quantity,
            item.work_price,
            item.mat_price,
            total_formula,
            source_display,
            ", ".join(flags) if flags else "",
        ])

        # Row highlight priority: scan error > AI-estimated > optimised > analogue
        if scan_err:
            fill = _SCAN_ERR_FILL
        elif is_estimated:
            fill = _ESTIMATED_FILL
        elif getattr(item, "is_optimized", False):
            fill = _OPT_FILL
        elif getattr(item, "is_analogue", False):
            fill = _ANALOGUE_FILL
        else:
            fill = None

        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).fill = fill

        if is_estimated:
            estimated_items.append(item)

    # Grand total row — formula over the entire Стоимость column
    ws.append([])
    last_data = ws.max_row - 1
    tr = ws.max_row + 1
    ws.cell(row=tr, column=4, value="ИТОГО").font = _BOLD
    ws.cell(row=tr, column=9, value=f"=SUM(I{data_start_row}:I{last_data})").font = _BOLD

    _set_estimate_col_widths(ws)

    # Sheet 2: AI-estimated items requiring review
    if estimated_items:
        we = wb.create_sheet("Требуют проверки")
        e_headers = ["№", "Раздел", "Наименование", "Ед.", "Кол-во", "Цена работ", "Цена мат.", "Источник", "Примечание"]
        for col, h in enumerate(e_headers, 1):
            cell = we.cell(row=1, column=col, value=h)
            cell.font = _BOLD
            cell.fill = _ESTIMATED_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for i, item in enumerate(estimated_items, 1):
            we.append([
                i, item.section, item.name, item.unit, item.quantity,
                item.work_price, item.mat_price,
                getattr(item, "source", None) or "",
                item.comment or "",
            ])
        for col, w in enumerate([5, 20, 50, 8, 8, 14, 14, 30, 50], 1):
            we.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _set_estimate_col_widths(ws) -> None:
    for col, w in enumerate([5, 20, 12, 50, 8, 8, 14, 14, 14, 30, 20], 1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ---------------------------------------------------------------------------
# Compare report Excel  (Модуль 2.9)
# ---------------------------------------------------------------------------

def build_compare_excel(diff: dict, analytics_text: str) -> bytes:
    """
    Build comparison report Excel (module 2.9).

    Sheet 1: «Расхождения» — structured diff table (code).
    Sheet 2: «Аналитика» — Claude's commentary on the diff.
    """
    wb = Workbook()

    # Sheet 1: diff table
    wd = wb.active
    wd.title = "Расхождения"
    d_headers = ["Тип расхождения", "Наименование", "Ед.", "Кол-во (Смета)", "Кол-во (Проект)", "Разница %", "Примечание"]
    for col, h in enumerate(d_headers, 1):
        cell = wd.cell(row=1, column=col, value=h)
        cell.font = _BOLD
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    _EXTRA_FILL    = PatternFill(fill_type="solid", fgColor="FFCDD2")   # red — in smeta, not in project
    _MISSING_FILL  = PatternFill(fill_type="solid", fgColor="C8E6C9")   # green — in project, not in smeta
    _QTY_FILL      = PatternFill(fill_type="solid", fgColor="FFF9C4")   # yellow — qty mismatch

    row = 2
    for entry in diff.get("extra", []):
        wd.append(["Лишняя позиция", entry.get("name", ""), entry.get("unit", ""),
                   entry.get("qty_smeta", ""), "", "", "Есть в смете, нет в проекте"])
        for c in range(1, len(d_headers) + 1):
            wd.cell(row=row, column=c).fill = _EXTRA_FILL
        row += 1

    for entry in diff.get("missing", []):
        wd.append(["Забытая позиция", entry.get("name", ""), entry.get("unit", ""),
                   "", entry.get("qty_project", ""), "", "Есть в проекте, нет в смете"])
        for c in range(1, len(d_headers) + 1):
            wd.cell(row=row, column=c).fill = _MISSING_FILL
        row += 1

    for entry in diff.get("qty_mismatch", []):
        qs = entry.get("qty_smeta", 0) or 0
        qp = entry.get("qty_project", 0) or 0
        pct = round((qs - qp) / qp * 100, 1) if qp else ""
        wd.append(["Расхождение объёма", entry.get("name", ""), entry.get("unit", ""),
                   qs, qp, pct, ""])
        for c in range(1, len(d_headers) + 1):
            wd.cell(row=row, column=c).fill = _QTY_FILL
        row += 1

    for col, w in enumerate([20, 50, 8, 14, 14, 10, 40], 1):
        wd.column_dimensions[get_column_letter(col)].width = w

    # Sheet 2: Claude analytics text
    wa = wb.create_sheet("Аналитика")
    wa.merge_cells("A1:G1")
    title_cell = wa.cell(row=1, column=1, value="Аналитический комментарий")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    for i, line in enumerate(analytics_text.splitlines(), start=2):
        wa.cell(row=i, column=1, value=line)
    wa.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Helpers (unchanged from original)
# ---------------------------------------------------------------------------

def build_kp_excel(items, comment: str = "") -> bytes:
    """Build commercial proposal request Excel for sending to suppliers."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Запрос КП"

    if comment:
        ws.merge_cells("A1:H1")
        c = ws.cell(row=1, column=1, value=f"Комментарий: {comment}")
        c.font = Font(italic=True, color="555555")
        ws.row_dimensions[1].height = 30
        start_row = 3
    else:
        start_row = 2

    headers = ["№", "Наименование", "Ед.изм.", "Кол-во", "Комментарий", "Цена поставщика", "Срок поставки", "Поставщик"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row - 1, column=col, value=h)
        cell.font = _BOLD
        cell.fill = PatternFill(fill_type="solid", fgColor="FFF9C4")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, item in enumerate(items, 1):
        row = [i, item.name, item.unit, item.quantity, getattr(item, "comment", "") or "", "", "", ""]
        ws.append(row)

    widths = [5, 50, 10, 10, 30, 16, 16, 25]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_separation_sheet_excel(items, title: str = "Разделительная ведомость") -> bytes:
    """Build separation sheet Excel: items grouped by section."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ведомость"

    ws.merge_cells("A1:J1")
    tc = ws.cell(row=1, column=1, value=title)
    tc.font = Font(bold=True, size=13)
    tc.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    from itertools import groupby
    sorted_items = sorted(items, key=lambda i: (i.section or ""))
    row = 2
    grand_total = 0.0

    for section, group in groupby(sorted_items, key=lambda i: i.section or ""):
        grp = list(group)
        ws.merge_cells(f"A{row}:J{row}")
        sh = ws.cell(row=row, column=1, value=section or "Без раздела")
        sh.font = Font(bold=True)
        sh.fill = _HEADER_FILL
        sh.alignment = Alignment(horizontal="left")
        row += 1

        for col, h in enumerate(["№", "Тип", "Наименование", "Ед.", "Кол-во", "Цена работ", "Цена мат.", "Стоимость работ", "Стоимость мат.", "Итого"], 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = _BOLD
            c.fill = PatternFill(fill_type="solid", fgColor="EEF2F9")
            c.alignment = Alignment(horizontal="center", wrap_text=True)
        row += 1

        sec_total = 0.0
        for j, item in enumerate(grp, 1):
            wp = item.work_price or 0
            mp = item.mat_price or 0
            qty = item.quantity or 0
            cost_w = wp * qty
            cost_m = mp * qty
            total = cost_w + cost_m
            sec_total += total
            for col, val in enumerate([j, item.type, item.name, item.unit, qty, wp, mp, cost_w, cost_m, total], 1):
                ws.cell(row=row, column=col, value=val)
            row += 1

        for col in range(1, 11):
            ws.cell(row=row, column=col).fill = _HEADER_FILL
        ws.cell(row=row, column=3, value="Итого по разделу").font = _BOLD
        ws.cell(row=row, column=10, value=sec_total).font = _BOLD
        grand_total += sec_total
        row += 2

    for col in range(1, 11):
        ws.cell(row=row, column=col).fill = PatternFill(fill_type="solid", fgColor="C5CAE9")
    ws.cell(row=row, column=3, value="ИТОГО").font = Font(bold=True, size=11)
    ws.cell(row=row, column=10, value=grand_total).font = Font(bold=True, size=11)

    widths = [5, 12, 50, 8, 8, 14, 14, 16, 16, 16]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_estimate_excel(data: bytes) -> list[dict]:
    """Parse uploaded Excel and return list of item dicts."""
    wb = load_workbook(io.BytesIO(data), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True)) if ws else []
    items = []
    for i, row in enumerate(rows):
        if not row or not any(row):
            continue
        try:
            name = str(row[3]).strip() if row[3] else (str(row[0]).strip() if row[0] else "")
            if not name or name.upper() == "ИТОГО":
                continue
            items.append({
                "position": i + 1,
                "section": str(row[1]).strip() if row[1] else "",
                "type": str(row[2]).strip() if row[2] else "Работа",
                "name": name,
                "unit": str(row[4]).strip() if row[4] else "шт",
                "quantity": float(row[5]) if row[5] else 1.0,
                "work_price": float(row[6]) if row[6] else 0.0,
                "mat_price": float(row[7]) if row[7] else 0.0,
                "source_url": str(row[9]).strip() if len(row) > 9 and row[9] else None,
            })
        except (ValueError, TypeError):
            continue
    return items
