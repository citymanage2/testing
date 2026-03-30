"""Generate KS-2 and KS-3 Excel documents (Russian construction act forms)."""
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter


def _thin_border(sides="all"):
    thin = Side(style="thin")
    none = Side(style=None)
    kwargs = {k: thin if sides == "all" else none for k in ("left", "right", "top", "bottom")}
    if sides != "all":
        for s in sides.split(","):
            kwargs[s.strip()] = thin
    return Border(**kwargs)


def _mc(ws, min_row, min_col, max_row, max_col, value="", bold=False, align="center", size=10, wrap=False, fill=None):
    """Merge cells, set value and style."""
    from openpyxl.utils import get_column_letter
    ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
    cell = ws.cell(row=min_row, column=min_col)
    cell.value = value
    cell.font = Font(name="Times New Roman", bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def build_ks2(
    items: list,          # list of EstimateItem ORM objects
    extras: dict,         # overhead_pct, overhead_sum, transport_pct, transport_sum, etc.
    company: dict | None, # {name, inn, address, ...} or None
    contractor: dict | None,  # {name, inn, address, ...} or None
    period_start: date,
    period_end: date,
    act_number: str,
    vat_rate: float = 20.0,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "КС-2"

    # Column widths (A=4, B=5, C=5, D=30, E=8, F=10, G=10, H=12, I=12, J=12)
    col_widths = [4, 5, 5, 30, 8, 10, 10, 12, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # Title
    _mc(ws, row, 1, row, 10, "АКТ О ПРИЁМКЕ ВЫПОЛНЕННЫХ РАБОТ", bold=True, size=13)
    row += 1
    _mc(ws, row, 1, row, 10, f"Форма № КС-2", size=10)
    row += 1
    ws.row_dimensions[row].height = 6
    row += 1

    # Meta block
    def meta_line(label, value, r):
        ws.cell(row=r, column=1, value=label).font = Font(name="Times New Roman", size=9)
        _mc(ws, r, 4, r, 10, value, align="left", size=9)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left")

    meta_line("Инвестор (Заказчик):", (contractor or {}).get("name", ""), row); row += 1
    meta_line("ИНН/КПП Заказчика:", f'{(contractor or {}).get("inn","")}/{(contractor or {}).get("kpp","")}', row); row += 1
    meta_line("Подрядчик:", (company or {}).get("name", ""), row); row += 1
    meta_line("ИНН/КПП Подрядчика:", f'{(company or {}).get("inn","")}/{(company or {}).get("kpp","")}', row); row += 1
    meta_line("Номер акта:", act_number, row); row += 1
    meta_line("Период выполнения работ:", f'с {period_start.strftime("%d.%m.%Y")} по {period_end.strftime("%d.%m.%Y")}', row); row += 1
    row += 1

    # Table header
    header_fill = "D9E1F2"
    headers = ["№ п/п", "№ позиции по смете", "Наименование работ и затрат", "Единица\nизмерения", "Кол-во", "Цена\nединицы", "Стоимость\nработ", "Стоимость\nматериалов", "Стоимость\nитого", "Примечание"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Times New Roman", bold=True, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=header_fill)
        c.border = _thin_border()
    ws.row_dimensions[row].height = 36
    row += 1

    # Number row
    for col in range(1, 11):
        c = ws.cell(row=row, column=col, value=col)
        c.font = Font(name="Times New Roman", size=9)
        c.alignment = Alignment(horizontal="center")
        c.border = _thin_border()
    row += 1

    total_work = 0.0
    total_mat = 0.0

    # Group by section
    current_section = None
    pos = 1
    for item in items:
        if getattr(item, "row_type", "item") == "section_header":
            # Section header row
            _mc(ws, row, 1, row, 10, item.name or "", bold=True, align="left", size=9, fill="EBF3FB")
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = _thin_border()
            row += 1
            continue

        work_total = round(item.work_price * item.quantity, 2)
        mat_total = round(item.mat_price * item.quantity, 2)
        item_total = round(work_total + mat_total, 2)
        total_work += work_total
        total_mat += mat_total

        vals = [pos, "", item.name, item.unit, item.quantity,
                item.work_price + item.mat_price, work_total, mat_total, item_total, ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name="Times New Roman", size=9)
            c.alignment = Alignment(horizontal="center" if col != 3 else "left", vertical="center", wrap_text=(col == 3))
            c.border = _thin_border()
            if col in (5, 6, 7, 8, 9) and isinstance(v, float):
                c.number_format = '#,##0.00'
        row += 1
        pos += 1

    # Extras (overhead, transport, contingency)
    base_total = total_work + total_mat
    overhead = round(extras.get("overhead_sum", 0) + base_total * extras.get("overhead_pct", 0) / 100, 2)
    transport = round(extras.get("transport_sum", 0) + base_total * extras.get("transport_pct", 0) / 100, 2)
    contingency = round(extras.get("contingency_sum", 0) + base_total * extras.get("contingency_pct", 0) / 100, 2)
    grand_base = base_total + overhead + transport + contingency
    vat = round(grand_base * vat_rate / 100, 2)
    grand_total = grand_base + vat

    totals = [
        ("Итого прямые затраты:", base_total),
        (f"Накладные расходы:", overhead),
        (f"Транспортные расходы:", transport),
        (f"Непредвиденные расходы:", contingency),
        (f"Итого без НДС:", grand_base),
        (f"НДС {vat_rate}%:", vat),
        ("ИТОГО с НДС:", grand_total),
    ]
    for label, amount in totals:
        _mc(ws, row, 1, row, 8, label, bold=("ИТОГО" in label), align="right", size=9)
        c = ws.cell(row=row, column=9, value=amount)
        c.font = Font(name="Times New Roman", bold=("ИТОГО" in label), size=9)
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal="right")
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = _thin_border()
        row += 1

    # Signatures
    row += 2
    ws.cell(row=row, column=1, value="Сдал:").font = Font(name="Times New Roman", size=10)
    ws.cell(row=row, column=6, value="Принял:").font = Font(name="Times New Roman", size=10)
    row += 2
    ws.cell(row=row, column=1, value="_______________________").font = Font(name="Times New Roman", size=10)
    ws.cell(row=row, column=6, value="_______________________").font = Font(name="Times New Roman", size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ks3(
    ks2_amount: float,
    company: dict | None,
    contractor: dict | None,
    period_start: date,
    period_end: date,
    act_number: str,
    vat_rate: float = 20.0,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "КС-3"

    col_widths = [6, 30, 15, 15, 15, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    _mc(ws, row, 1, row, 6, "СПРАВКА О СТОИМОСТИ ВЫПОЛНЕННЫХ РАБОТ И ЗАТРАТ", bold=True, size=13)
    row += 1
    _mc(ws, row, 1, row, 6, "Форма № КС-3", size=10)
    row += 1; row += 1

    def meta_line(label, value, r):
        ws.cell(row=r, column=1, value=label).font = Font(name="Times New Roman", size=9)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        c = ws.cell(row=r, column=2, value=value)
        c.font = Font(name="Times New Roman", size=9)
        c.alignment = Alignment(horizontal="left")

    meta_line("Заказчик:", (contractor or {}).get("name", ""), row); row += 1
    meta_line("Подрядчик:", (company or {}).get("name", ""), row); row += 1
    meta_line("Номер справки:", act_number, row); row += 1
    meta_line("Отчётный период:", f'{period_start.strftime("%d.%m.%Y")} — {period_end.strftime("%d.%m.%Y")}', row); row += 1
    row += 1

    fill = "D9E1F2"
    hdrs = ["№", "Наименование стройки,\nобъекта, вида работ", "Стоимость\nработ с начала строительства", "Стоимость\nс начала года", "Стоимость\nза отчётный период", "Стоимость\nмат-лов и конструкций"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Times New Roman", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[row].height = 40
    row += 1

    # Data row
    vat = round(ks2_amount * vat_rate / 100, 2)
    grand = ks2_amount + vat
    data = [1, "Выполнение работ по договору подряда", grand, grand, grand, ""]
    for col, v in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(name="Times New Roman", size=9)
        c.border = _thin_border()
        if isinstance(v, float):
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right")
        else:
            c.alignment = Alignment(horizontal="left" if col == 2 else "center")
    row += 1

    # VAT row
    vat_row = ["", f"В том числе НДС {vat_rate}%:", vat, vat, vat, ""]
    for col, v in enumerate(vat_row, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(name="Times New Roman", italic=True, size=9)
        c.border = _thin_border()
        if isinstance(v, float):
            c.number_format = '#,##0.00'
            c.alignment = Alignment(horizontal="right")
        else:
            c.alignment = Alignment(horizontal="left" if col == 2 else "center")
    row += 2

    ws.cell(row=row, column=1, value="Сдал:").font = Font(name="Times New Roman", size=10)
    ws.cell(row=row, column=4, value="Принял:").font = Font(name="Times New Roman", size=10)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_estimate_xlsx(
    items: list,
    extras: dict,
    company: dict | None,
    title: str,
    vat_rate: float = 20.0,
) -> bytes:
    """Professional formatted estimate as Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"
    ws.page_setup.orientation = "landscape"

    col_widths = [5, 12, 12, 35, 8, 10, 12, 12, 14, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1
    _mc(ws, row, 1, row, 11, (company or {}).get("name", ""), align="left", size=10)
    if (company or {}).get("inn"):
        row += 1
        _mc(ws, row, 1, row, 11, f'ИНН: {company["inn"]}  КПП: {company.get("kpp","")}  Адрес: {company.get("address","")}', align="left", size=9)
    row += 1
    _mc(ws, row, 1, row, 11, title, bold=True, size=14)
    row += 1; row += 1

    fill = "D9E1F2"
    hdrs = ["№", "Раздел", "Тип", "Наименование", "Ед.", "Кол-во", "Цена\nработ", "Цена\nмат.", "Стоимость", "Источник", "Комментарий"]
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Times New Roman", bold=True, size=9)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thin_border()
    ws.row_dimensions[row].height = 28
    row += 1

    total_work = 0.0
    total_mat = 0.0
    pos = 1
    for item in items:
        if getattr(item, "row_type", "item") == "section_header":
            _mc(ws, row, 1, row, 11, item.name or "", bold=True, align="left", size=9, fill="EBF3FB")
            for col in range(1, 12):
                ws.cell(row=row, column=col).border = _thin_border()
            row += 1
            continue

        total = round((item.work_price + item.mat_price) * item.quantity, 2)
        total_work += round(item.work_price * item.quantity, 2)
        total_mat += round(item.mat_price * item.quantity, 2)

        vals = [pos, item.section, item.type, item.name, item.unit, item.quantity,
                item.work_price, item.mat_price, total, item.source_url or "", item.comment or ""]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=col, value=v)
            c.font = Font(name="Times New Roman", size=9)
            c.border = _thin_border()
            c.alignment = Alignment(horizontal="left" if col in (3, 4, 10, 11) else "center", vertical="top", wrap_text=(col in (4, 10, 11)))
            if col in (6, 7, 8, 9) and isinstance(v, float):
                c.number_format = '#,##0.00'
        row += 1
        pos += 1

    base = total_work + total_mat
    overhead = round(extras.get("overhead_sum", 0) + base * extras.get("overhead_pct", 0) / 100, 2)
    transport = round(extras.get("transport_sum", 0) + base * extras.get("transport_pct", 0) / 100, 2)
    contingency = round(extras.get("contingency_sum", 0) + base * extras.get("contingency_pct", 0) / 100, 2)
    grand_base = base + overhead + transport + contingency
    vat = round(grand_base * vat_rate / 100, 2)
    grand = grand_base + vat

    row += 1
    summary = [
        ("Работы:", total_work), ("Материалы:", total_mat), ("Итого:", base),
        ("Накладные расходы:", overhead), ("Транспортные расходы:", transport),
        ("Непредвиденные расходы:", contingency), ("Итого без НДС:", grand_base),
        (f"НДС {vat_rate}%:", vat), ("ИТОГО с НДС:", grand),
    ]
    for label, amount in summary:
        _mc(ws, row, 1, row, 8, label, bold=("ИТОГО" in label), align="right", size=9)
        c = ws.cell(row=row, column=9, value=amount)
        c.font = Font(name="Times New Roman", bold=("ИТОГО" in label), size=10)
        c.number_format = '#,##0.00'
        c.alignment = Alignment(horizontal="right")
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = _thin_border()
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
